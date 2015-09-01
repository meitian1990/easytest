__author__ = 'min.sun'

from openpyxl import *
#获得excel中的测试套件
def getsuit(filename):
    testsuitlist = []
    try:
        wb = load_workbook(filename = filename)
        ws = wb.get_sheet_by_name("测试套件")
        #ws.get_highest_row()获得excel里面的行数，因为range(m.n)是取m-n的数据，不包含n，所以要在行数上+1，
        # 而excel的cell是从（1,1）开始算的，第一行的数据不要，所以range是从2开始的。
        #本来想通过遍历行和列，然后存储数据的，但是考虑到存储的字典名不同，所以还是每个表写一个吧
        for i in range(2,ws.get_highest_row()+1):
            if(ws.cell(row=i,column=3).value == "YES"):
                list = {}
                list["suitname"] = ws.cell(row =i ,column= 1).value
                list["suitmethodname"] = ws.cell(row=i,column=2).value
                list["suitisperform"] = ws.cell(row=i,column=3).value
                testsuitlist.append(list)
        # print(testsuitlist)
    except IOError:
        print("file:"+filename+" is not exesit")
    return testsuitlist
#获得excel中的XX用例模块，用suit作为模块的sheet名称，例如登录用例
def gettestcase(filename,suit):
    testcaselist = []
    try:
        wb = load_workbook(filename = filename)
        try:
            ws = wb.get_sheet_by_name(suit)
            for i in range(2,ws.get_highest_row()+1):
                list = {}
                list["teststeps"] = ws.cell(row =i ,column= 1).value
                list["testdescribe"] = ws.cell(row=i,column=2).value
                list["pageobject"] = ws.cell(row=i,column=3).value
                list["testobject"] = ws.cell(row=i,column=4).value
                list["testoperation"] = ws.cell(row=i,column=5).value
                list["testdata"] = ws.cell(row=i,column=6).value
                list["testcheck"] = ws.cell(row=i,column=7).value
                testcaselist.append(list)
            return testcaselist
        except:
            print("请先在用例中增加对应的用例模块：",suit)
    except IOError:
        print("file:"+filename+" is not exesit")
#获得excel中的测试元素
def gettestelement(filename):
    testelementlist = []
    try:
        wb = load_workbook(filename = filename)
        ws = wb.get_sheet_by_name("测试元素")
        for i in range(2,ws.get_highest_row()+1):
            list = {}
            list["pageobject"] = ws.cell(row =i ,column= 1).value
            list["testobject"] = ws.cell(row=i,column=2).value
            list["positionway"] = ws.cell(row=i,column=3).value
            list["testelement"] = ws.cell(row=i,column=4).value
            testelementlist.append(list)
        return testelementlist
    except IOError:
        print("file:"+filename+" is not exesit")

#测试用例，通过匹配用例模块和元素来找到用例对应的元素
def testcase(filename,suit):
    #testcaselistall用来存储匹配后的用例和对应的元素。
    testcaselistall = []
    testcaselist = gettestcase(filename,suit)
    testelementlist = gettestelement(filename)
    #遍历用例模块和测试元素sheet，匹配出与用例模块相等的测试元素，保存结果
    #加上testcaselist的if判断是否为空，是为了防止没有找到对应模块用例的sheet，返回了一个空的list，如果为空不加if会报错的。
    if testcaselist is not None:
        for testcase in testcaselist:
            #加这个if语句是因为有的用例，assertTitle之类的是没有对应的测试元素的，如果直接走下面的for循环，该用例不会加到测试组里
            if testcase["testoperation"] == "assertTitle":
                list = {}
                list["pageobject"] = testcase["pageobject"]
                list["testobject"] = testcase["testobject"]
                # list["positionway"] = testelement["positionway"]
                # list["testelement"] = testelement["testelement"]
                list["testoperation"] = testcase["testoperation"]
                list["testdata"] = testcase["testdata"]
                list["testcheck"] = testcase["testcheck"]
                testcaselistall.append(list)
            for testelement in testelementlist:
                if testcase["testobject"] == testelement["testobject"]:
                    list = {}
                    list["pageobject"] = testcase["pageobject"]
                    list["testobject"] = testcase["testobject"]
                    list["positionway"] = testelement["positionway"]
                    list["testelement"] = testelement["testelement"]
                    list["testoperation"] = testcase["testoperation"]
                    list["testdata"] = testcase["testdata"]
                    list["testcheck"] = testcase["testcheck"]
                    testcaselistall.append(list)
        return testcaselistall

# filename = "C:/Users/min.sun/Desktop/自动化测试/testcase.xlsx"

#测试一下结果
# a =getsuit(filename)
# print(a)
# b = gettestcase(filename,"登录用例")
# print(b)
# c = gettestelement(filename)
# print(c)
# print(testcase(filename,"登录用例"))